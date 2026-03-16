#!/usr/bin/env python3
"""
Chuyển đổi MES_TAG_NAMING_STANDARD.md sang file Excel
"""
import re
import os

def parse_markdown_tables(content):
    """Trích xuất các bảng markdown từ nội dung"""
    tables = []
    lines = content.split('\n')
    i = 0
    current_section = "Overview"
    current_subsection = ""
    
    while i < len(lines):
        line = lines[i]
        
        # Cập nhật section từ heading
        if line.startswith('## ') and not line.startswith('### '):
            current_section = line[3:].strip()
            current_subsection = ""
        elif line.startswith('### '):
            current_subsection = line[4:].strip()
        
        # Bắt đầu bảng markdown (dòng bắt đầu bằng |)
        if line.strip().startswith('|') and line.strip().endswith('|'):
            table_lines = []
            table_lines.append(line)
            i += 1
            
            while i < len(lines) and lines[i].strip().startswith('|') and lines[i].strip().endswith('|'):
                table_lines.append(lines[i])
                i += 1
            
            # Parse bảng
            if len(table_lines) >= 2:  # Có header và ít nhất 1 dòng separator
                rows = []
                for tl in table_lines:
                    # Bỏ qua dòng separator (|---|---|)
                    cells = [c.strip() for c in tl.strip().split('|')[1:-1]]
                    if cells and not all(re.match(r'^[-:]+$', c.replace(' ', '')) for c in cells):
                        rows.append(cells)
                
                if rows:
                    tables.append({
                        'section': current_section,
                        'subsection': current_subsection,
                        'data': rows
                    })
            continue
        
        i += 1
    
    return tables

def strip_backticks(text):
    """Bỏ dấu backtick trong giá trị khi xuất Excel"""
    if not isinstance(text, str):
        return text
    return text.replace('`', '')

def sanitize_sheet_name(name):
    """Tên sheet Excel tối đa 31 ký tự, không chứa: \\ / * ? : [ ]"""
    invalid = ['\\', '/', '*', '?', ':', '[', ']']
    for c in invalid:
        name = name.replace(c, '_')
    return name[:31]

def main():
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("Installing openpyxl...")
        os.system("pip install openpyxl -q")
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Đường dẫn file
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)
    md_path = os.path.join(project_dir, 'docs', 'MES_TAG_NAMING_STANDARD.md')
    output_path = os.path.join(project_dir, 'docs', 'MES_TAG_NAMING_STANDARD.xlsx')
    
    if not os.path.exists(md_path):
        print(f"Không tìm thấy file: {md_path}")
        return
    
    # Đọc markdown
    with open(md_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    tables = parse_markdown_tables(content)
    
    # Tạo workbook
    wb = openpyxl.Workbook()
    
    # Xóa sheet mặc định nếu có nhiều bảng
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Gom nhóm bảng theo section
    section_tables = {}
    for t in tables:
        sec = t['section']
        if sec not in section_tables:
            section_tables[sec] = []
        section_tables[sec].append({
            'subsection': t.get('subsection', ''),
            'data': t['data']
        })
    
    sheet_count = 0
    for section, table_list in section_tables.items():
        sheet_name = sanitize_sheet_name(section)
        # Tránh trùng tên sheet
        base_name = sheet_name
        counter = 1
        while sheet_name in [s.title for s in wb.worksheets]:
            sheet_name = f"{base_name[:28]}_{counter}"
            counter += 1
        
        ws = wb.create_sheet(title=sheet_name)
        row_num = 1
        
        for table_idx, item in enumerate(table_list):
            table_data = item['data']
            subsection = item.get('subsection', '')
            
            if table_idx > 0:
                row_num += 2  # Khoảng cách giữa các bảng
            
            # Thêm tiêu đề subsection nếu có
            if subsection:
                ws.cell(row=row_num, column=1, value=strip_backticks(subsection))
                ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
                row_num += 1
            
            for r_idx, row in enumerate(table_data):
                for c_idx, cell_val in enumerate(row):
                    cell_val_clean = strip_backticks(cell_val)
                    cell = ws.cell(row=row_num + r_idx, column=c_idx + 1, value=cell_val_clean)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if r_idx == 0:  # Header
                        cell.font = header_font_white
                        cell.fill = header_fill
                
                # Auto width
                for c_idx in range(len(row)):
                    col_letter = get_column_letter(c_idx + 1)
                    current_width = ws.column_dimensions[col_letter].width or 10
                    cell_len = min(len(str(row[c_idx])) + 2, 50)
                    ws.column_dimensions[col_letter].width = max(current_width, cell_len)
            
            row_num += len(table_data)
        
        sheet_count += 1
    
    # Lưu file
    wb.save(output_path)
    print(f"Exported: {output_path}")
    print(f"Sheets: {sheet_count}")
    return output_path

if __name__ == '__main__':
    main()
