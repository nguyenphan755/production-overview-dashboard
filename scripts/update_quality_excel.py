#!/usr/bin/env python3
"""
Cập nhật file MES_TAG_Quality.xlsx: thêm/cập nhật sheet "3.5 Tagname Chat luong"
với bảng tagname rút gọn từ docs/MES_TAG_NAMING_STANDARD.md mục 3.5.
"""
import os
import sys

def main():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Installing openpyxl...")
        os.system(f'"{sys.executable}" -m pip install openpyxl -q')
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)
    # File gốc từ nhà máy (Downloads) hoặc trong project
    for base in [os.path.join(os.path.expanduser("~"), "Downloads"), project_dir]:
        path = os.path.join(base, "MES_TAG_Quality.xlsx")
        if os.path.exists(path):
            excel_path = path
            break
    else:
        # Tạo mới trong project
        excel_path = os.path.join(project_dir, "MES_TAG_Quality.xlsx")
        print(f"Không tìm thấy MES_TAG_Quality.xlsx, sẽ tạo mới: {excel_path}")

    # Dữ liệu mục 3.5 (Tagname rút gọn) - đồng bộ với MES_TAG_NAMING_STANDARD.md
    header = ["Công đoạn", "Tham số chất lượng", "Tagname (đề xuất)", "Đơn vị", "Thiết bị / Nguồn", "Ghi chú"]
    rows = [
        ["Raw Material", "Điện trở suất 20°C", "Q_Raw_Rho20", "Ω·mm²/m", "Micro-ohmmeter, QMS", "Cầu đo DO5000"],
        ["Raw Material", "Đường kính", "Q_Raw_Dia", "mm", "Laser micrometer", "Panme điện tử"],
        ["Raw Material", "Bề mặt", "Q_Raw_Sur", "-", "Camera / Visual", "Xem xét thêm TB Machine Vision"],
        ["Raw Material", "Độ bền kéo", "Q_Raw_Tens", "Mpa", "UTM", "Máy kháng kéo"],
        ["Raw Material", "Độ giãn dài", "Q_Raw_Elong", "%", "UTM", "-"],
        ["Raw Material", "Tỷ trọng", "Q_Raw_Dens", "g/cm³", "Density balance", "Phòng KCS NM CADIVI Khí cụ điện"],
        ["Raw Material", "Độ cứng Shore", "Q_Raw_Shore", "A/D", "Shore Durometer", "-"],
        ["Drawing", "Đường kính dây", "Q_Drw_WDia", "mm", "Laser micrometer, Online", "Keyence IG-20/LS5000"],
        ["Drawing", "Sai lệch đường kính", "Q_Drw_DiaDev", "mm", "MES SPC", "Nhập dung sai trên HMI link PLC"],
        ["Drawing", "Số lần đứt dây/ca", "Q_Drw_WBreak", "-", "Machine sensor", "Counter đưa về PLC, cảm biến báo đứt"],
        ["Drawing", "Điện trở DC ở 20°C", "Q_Drw_Rdc20", "Ω/km", "Online resistance meter", "Xem xét thêm TB đo Resistor online"],
        ["Drawing", "Độ bền kéo", "Q_Drw_Tens", "Mpa", "UTM, QMS", "Máy kháng kéo"],
        ["Drawing", "Độ giãn dài", "Q_Drw_Elong", "%", "UTM, QMS", "-"],
        ["Annealing", "Điện trở DC ở 20°C", "Q_Ann_Rdc20", "Ω/km", "Online resistance meter", "Xem xét thêm TB đo Resistor online"],
        ["Annealing", "Độ bền kéo", "Q_Ann_Tens", "Mpa", "UTM, QMS", "Máy kháng kéo"],
        ["Annealing", "Độ giãn dài", "Q_Ann_Elong", "%", "UTM", "-"],
        ["Annealing", "Đường kính dây", "Q_Ann_WDia", "mm", "Laser micrometer, Online", "Keyence IG-20/LS5000"],
        ["Stranding", "Đường kính ruột dẫn", "Q_Str_CondOD", "mm", "Laser OD gauge, Online", "Keyence IG-20/LS5000"],
        ["Stranding", "Điện trở DC ở 20°C", "Q_Str_Rdc20", "Ω/km", "Online resistance meter", "Xem xét thêm TB đo Resistor online"],
        ["Stranding", "Bước xoắn", "Q_Str_Lay", "mm", "Encoder calculation", "-"],
        ["Stranding", "Hướng xoắn (S/Z)", "Q_Str_Dir", "spec", "Recipe control", "-"],
        ["Stranding", "Số sợi", "Q_Str_NStr", "-", "Manual verify", "Đếm số lượng Pay-Off Active"],
        ["Mica Armouring", "Số lớp mica", "Q_Mic_MLay", "-", "PLC + Recipe", "Mặc định là 2"],
        ["Mica Armouring", "Độ dày băng", "Q_Mic_TThick", "mm", "Micrometer", "Nhập trên HMI link PLC"],
        ["Mica Armouring", "Chiều rộng băng", "Q_Mic_TWid", "mm", "Caliper", "-"],
        ["Mica Armouring", "Độ chồng mí", "Q_Mic_Overlap", "%", "Encoder / Vision", "-"],
        ["Mica Armouring", "Hướng giáp (S/Z)", "Q_Mic_Dir", "-", "Direction sensor", "-"],
        ["Mica Armouring", "Tốc độ line", "Q_Mic_Speed", "m/min", "Line encoder + PLC", "-"],
        ["Insulating", "Đường kính ngoài", "Q_Ins_OD", "mm", "Laser OD gauge, Online", "Keyence IG-20/LS5000"],
        ["Insulating", "Độ dày cách điện (Avg./Min)", "Q_Ins_InsTh", "mm", "X-ray / Ultrasonic", "Xem xét thêm TB đo realtime"],
        ["Insulating", "Độ lệch tâm", "Q_Ins_Ecc", "%", "X-ray gauge", "-"],
        ["Insulating", "Bề mặt", "Q_Ins_Sur", "-", "Camera / Visual", "Xem xét thêm TB Machine Vision"],
        ["Insulating", "Spark test", "Q_Ins_Spark", "kV", "Spark tester", "Xem xét nâng cấp thiết bị có truyền thông"],
        ["Insulating", "Điện trở cách điện", "Q_Ins_IR", "MΩ·km", "IR tester, QMS", "Xem xét thêm TB đo Resistor online"],
        ["Insulating", "Độ bền kéo", "Q_Ins_Tens", "Mpa", "UTM, QMS", "Máy kháng kéo"],
        ["Insulating", "Độ giãn dài", "Q_Ins_Elong", "%", "UTM, QMS", "-"],
        ["Cabling", "Số lõi ghép", "Q_Cab_Cores", "-", "Setup verify", "Nhập từ Tablet"],
        ["Cabling", "Đường kính tổng sau ghép", "Q_Cab_ODcab", "mm", "Laser OD gauge, QMS", "-"],
        ["Cabling", "Bước ghép", "Q_Cab_Lay", "mm", "Encoder calculation", "-"],
        ["Cabling", "Hướng ghép (S/Z)", "Q_Cab_Dir", "spec", "Recipe control", "-"],
        ["Inner Sheathing", "Bề mặt", "Q_InSh_Sur", "-", "Camera / Visual", "Xem xét thêm TB Machine Vision"],
        ["Inner Sheathing", "Đường kính ngoài", "Q_InSh_OD", "mm", "Laser OD gauge, Online", "Keyence IG-20/LS5000"],
        ["Inner Sheathing", "Độ dày vỏ trong", "Q_InSh_InShTh", "mm", "X-ray / Ultrasonic", "Xem xét thêm TB Ultrasonic online"],
        ["Inner Sheathing", "Độ bền kéo", "Q_InSh_Tens", "Mpa", "UTM", "Máy kháng kéo"],
        ["Inner Sheathing", "Độ giãn dài", "Q_InSh_Elong", "%", "UTM", "-"],
        ["Armouring", "Bề mặt", "Q_Arm_Sur", "-", "Camera / Visual", "Xem xét thêm TB Machine Vision"],
        ["Armouring", "Độ dày", "Q_Arm_Thick", "mm", "Micrometer", "Nhập từ Tablet"],
        ["Armouring", "Chiều rộng", "Q_Arm_Wid", "mm", "Caliper", "-"],
        ["Armouring", "Độ chồng mí (%)", "Q_Arm_Overlap", "%", "Encoder / PLC", "-"],
        ["Armouring", "Độ bền kéo", "Q_Arm_Tens", "Mpa", "UTM, QMS", "Máy kháng kéo"],
        ["Armouring", "Đường kính ngoài sau giáp", "Q_Arm_ODarm", "mm", "Laser OD gauge, Online", "Keyence IG-20/LS5000"],
        ["Outer Sheathing", "Đường kính ngoài", "Q_OuSh_OD", "mm", "Laser OD gauge, Online", "Keyence IG-20/LS5000"],
        ["Outer Sheathing", "Độ dày vỏ (Avg./Min)", "Q_OuSh_ShTh", "mm", "X-ray / Ultrasonic", "Nhập từ Tablet"],
        ["Outer Sheathing", "Spark test", "Q_OuSh_Spark", "kV", "Spark tester", "-"],
        ["Outer Sheathing", "Trọng lượng", "Q_OuSh_Wt", "kg/km", "Inline scale", "-"],
        ["Outer Sheathing", "Độ bền kéo", "Q_OuSh_Tens", "Mpa", "UTM", "Máy kháng kéo"],
        ["Outer Sheathing", "Độ giãn dài", "Q_OuSh_Elong", "%", "UTM", "-"],
        ["Outer Sheathing", "Chữ in", "Q_OuSh_Print", "m", "Vision camera", "Xem xét thêm TB Machine Vision"],
        ["Final Testing", "Độ dày vỏ bọc (Avg./Min)", "Q_Fin_ShTh", "mm", "Microscope, QMS", "-"],
        ["Final Testing", "Độ dày cách điện (Avg./Min)", "Q_Fin_InsTh", "mm", "Microscope, QMS", "-"],
        ["Final Testing", "Điện trở DC ở 20°C", "Q_Fin_Rdc20", "Ω/km", "Online resistance meter, QMS", "-"],
        ["Final Testing", "Độ bền kéo", "Q_Fin_Tens", "Mpa", "UTM, QMS", "-"],
        ["Final Testing", "Độ giãn dài", "Q_Fin_Elong", "%", "UTM, QMS", "-"],
        ["Final Testing", "Điện trở cách điện", "Q_Fin_IR", "MΩ·km", "IR tester, QMS", "-"],
        ["Final Testing", "Thử điện áp HV", "Q_Fin_HV", "kV", "PD-HV system, QMS", "-"],
        ["Final Testing", "Thử phóng điện cục bộ PD", "Q_Fin_PD", "pC", "PD-HV system, QMS", "-"],
        ["Final Testing", "Đường kính tổng", "Q_Fin_OD", "mm", "Trên mẫu, QMS", "-"],
    ]

    sheet_name = "3.5 Tagname Chat luong"
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = wb.create_sheet(title=sheet_name[:31])
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for c, val in enumerate(header, 1):
        cell = ws.cell(row=1, column=c, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    for c in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    wb.save(excel_path)
    print("Updated sheet:", sheet_name)
    print("File:", excel_path)
    return excel_path

if __name__ == '__main__':
    main()
