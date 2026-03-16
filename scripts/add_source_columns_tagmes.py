#!/usr/bin/env python3
# Add columns "Kieu du lieu" and "Nguon du lieu" to c:\Users\Admin\Downloads\tagmes.xlsx
# Uses mapping from MES_TAG_TABLET_MES_SAP_SOURCE.md - normalized tag names

import re
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    import os
    os.system("pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import PatternFill

EXCEL_PATH = Path(r"c:\Users\Admin\Downloads\tagmes.xlsx")
# New file: same content + 2 columns (do not overwrite original)
OUTPUT_PATH = Path(r"c:\Users\Admin\Downloads\tagmes_with_source.xlsx")

# Dòng đề xuất bổ sung (Tablet/MES/SAP) - sẽ append vào cuối Excel và tô màu vàng
# (tagname_col_b, mo_ta_col_d, type, source)
NEW_SUGGESTED_ROWS = [
    ("Order_SapOrderId", "Mã lệnh sản xuất SAP", "VARCHAR(100)", "SAP->MES"),
    ("Order_ConfirmationStatus", "Trạng thái xác nhận sản xuất", "VARCHAR(50)", "MES->SAP; SAP->MES"),
    ("Order_SyncToSapAt", "Thời điểm đồng bộ lên SAP", "TIMESTAMPTZ", "MES"),
    ("Order_SyncToSapStatus", "success, pending, failed", "VARCHAR(50)", "MES"),
    ("Order_ScrapToSap", "Báo phế phẩm lên SAP", "BOOLEAN/TEXT", "MES->SAP"),
    ("{Order}.MaterialBom", "BOM định mức vật tư", "TEXT/JSON", "SAP->MES"),
    ("{Machine}.CurrentShiftStart", "Bắt đầu ca hiện tại", "TIMESTAMPTZ", "MES"),
    ("{Machine}.CurrentShiftEnd", "Kết thúc ca hiện tại", "TIMESTAMPTZ", "MES"),
    ("{Machine}.CoilLength", "Chiều dài cuộn (m)", "DECIMAL(12,2)", "PLC->API; MES"),
    ("{Machine}.CoilWeight", "Khối lượng cuộn (kg)", "DECIMAL(12,3)", "PLC->API; Tablet->MES"),
    ("Downtime_Reason", "Nguyên nhân dừng máy (mã)", "VARCHAR(100)", "Tablet->MES"),
    ("event_time", "Thời điểm sự kiện", "TIMESTAMPTZ", "Tablet->MES; PLC->API"),
    ("entered_by", "Người nhập (user đăng nhập)", "VARCHAR(255)", "Tablet->MES"),
    ("Order_RealTime", "Thời gian thực (hiển thị)", "TIMESTAMPTZ", "Tablet->MES"),
    ("Order_ORScanTime", "Thời điểm quét OR", "TIMESTAMPTZ", "Tablet->MES"),
    ("Order_ORScanBy", "Người quét OR", "VARCHAR(255)", "Tablet->MES"),
]
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Normalize tag from Excel: (Machine).X -> {Machine}_X, remove backticks
def norm(s):
    if s is None or (isinstance(s, float) and str(s) == "nan"):
        return ""
    s = str(s).strip().replace("`", "")
    s = re.sub(r"\(Machine\)\.", "{Machine}_", s, flags=re.I)
    s = re.sub(r"\{Machine\}\.", "{Machine}_", s, flags=re.I)
    s = re.sub(r"\(Order\)\.", "Order_", s, flags=re.I)
    s = re.sub(r"\{Order\}\.", "Order_", s, flags=re.I)
    s = re.sub(r"^Order\.", "Order_", s)
    s = re.sub(r"^Shift\.", "Shift_", s)
    s = re.sub(r"^Event\.", "Event_", s)
    s = re.sub(r"Q\.Drw\.", "Q.Dw.", s)
    s = re.sub(r"Q\.Dw\.", "Q_Drw_", s)
    s = re.sub(r"T\.Draw\.", "T.Dw.", s)
    return s

# Build lookup: tag (normalized) -> (data_type, source)
def build_lookup():
    d = {}
    # From MES_TAG_TABLET_MES_SAP_SOURCE.md - key = normalized tag, value = (type, source)
    pairs = [
        ("Tablet.ID", "VARCHAR(50)", "Tablet"),
        ("{Machine}_OrderId", "VARCHAR(100)", "SAP->MES; PLC->API"),
        ("{Machine}_OrderName", "VARCHAR(255)", "SAP->MES; Tablet->MES"),
        ("{Machine}_OrderCode", "VARCHAR(100)", "SAP->MES"),
        ("{Machine}_ProductName", "VARCHAR(255)", "SAP->MES; PLC->API"),
        ("{Machine}_MaterialCode", "VARCHAR(50)", "SAP->MES; PLC->API"),
        ("{Machine}_CustomerName", "VARCHAR(255)", "SAP->MES"),
        ("Order_StartTime", "TIMESTAMPTZ", "Tablet->MES"),
        ("Order_EndTime", "TIMESTAMPTZ", "Tablet->MES"),
        ("Order_RealTime", "TIMESTAMPTZ", "Tablet->MES"),
        ("Order_Status", "VARCHAR(50)", "Tablet->MES; PLC->API; MES->SAP"),
        ("Order_ProducedLength", "DECIMAL(12,2)", "MES"),
        ("Order_TargetLength", "DECIMAL(12,2)", "SAP->MES; Tablet->MES"),
        ("Order_Duration", "VARCHAR(50)", "MES"),
        ("Order_RemainTime", "VARCHAR(50)", "MES"),
        ("Order_EstTime", "TIMESTAMPTZ", "MES"),
        ("Order_FullTime", "TIMESTAMPTZ", "MES"),
        ("Order_MachineName", "VARCHAR(100)", "Tablet->MES; MES"),
        ("Order_QRScanTime", "TIMESTAMPTZ", "Tablet->MES"),
        ("Order_ORScanTime", "TIMESTAMPTZ", "Tablet->MES"),
        ("Order_QRScanBy", "VARCHAR(255)", "Tablet->MES"),
        ("Order_ORScanBy", "VARCHAR(255)", "Tablet->MES"),
        ("Order_PlannedLength", "DECIMAL(12,2)", "SAP->MES; Tablet->MES"),
        ("{Machine}_Status", "VARCHAR(50)", "PLC->API"),
        ("{Machine}_LineSpeed", "DECIMAL(10,2)", "PLC->API"),
        ("{Machine}_TargetSpeed", "DECIMAL(10,2)", "PLC->API; Tablet->MES"),
        ("{Machine}_LengthCounter", "DECIMAL(12,2)", "PLC->API"),
        ("{Machine}_LengthCounterLast", "DECIMAL(12,2)", "MES"),
        ("{Machine}_LengthCounterLastAt", "TIMESTAMPTZ", "MES"),
        ("{Machine}_LengthCounterDelta", "DECIMAL(12,2)", "MES"),
        ("{Machine}_LengthCounterReal", "TIMESTAMPTZ", "MES"),
        ("{Machine}_ProducedLength", "DECIMAL(12,2)", "MES"),
        ("{Machine}_TargetLength", "DECIMAL(12,2)", "SAP->MES; Tablet->MES"),
        ("{Machine}_CurrentShiftId", "VARCHAR(50)", "MES"),
        ("{Machine}_ShiftID", "VARCHAR(50)", "MES"),
        ("{Machine}_OperatorMain", "VARCHAR(255)", "Tablet->MES"),
        ("{Machine}_OperatorAux", "VARCHAR(255)", "Tablet->MES"),
        ("{Machine}_OperatorName", "VARCHAR(255)", "Tablet->MES"),
        ("Shift_HandoverTime", "TIMESTAMPTZ", "Tablet->MES"),
        ("Shift_HandoverBy", "VARCHAR(255)", "Tablet->MES"),
        ("Shift_ReceivedBy", "VARCHAR(255)", "Tablet->MES"),
        ("Shift_ReceiveBy", "VARCHAR(255)", "Tablet->MES"),
        ("Shift_HandoverNotes", "TEXT", "Tablet->MES"),
        ("{Machine}_BatchNumber", "VARCHAR(50)", "Tablet->MES; SAP->MES"),
        ("{Machine}_BobbinId", "VARCHAR(50)", "Tablet->MES; PLC->API"),
        ("{Machine}_BobbinID", "VARCHAR(50)", "Tablet->MES; PLC->API"),
        ("{Machine}_BobbinType", "VARCHAR(50)", "Tablet->MES"),
        ("{Machine}_BobbinWeight", "DECIMAL(12,3)", "Tablet->MES; PLC->API"),
        ("{Machine}_ProducedLengthOK", "DECIMAL(12,2)", "PLC->API; MES"),
        ("{Machine}_ProducedLengthNG", "DECIMAL(12,2)", "PLC->API; Tablet->MES"),
        ("{Machine}_PktIn", "VARCHAR(50)", "Tablet->MES; PLC->API"),
        ("{Machine}_PktInID", "VARCHAR(50)", "Tablet->MES; PLC->API"),
        ("{Machine}_PktOut", "VARCHAR(50)", "Tablet->MES; PLC->API"),
        ("{Machine}_PktOutID", "VARCHAR(50)", "Tablet->MES; PLC->API"),
        ("{Machine}_WeightNet", "DECIMAL(12,3)", "Tablet->MES; MES->SAP"),
        ("{Machine}_WeightGross", "DECIMAL(12,3)", "Tablet->MES; MES->SAP"),
        ("{Machine}_SampleCode", "VARCHAR(50)", "Tablet->MES; SAP->MES"),
        ("{Machine}_Notes", "TEXT", "Tablet->MES"),
        ("{Machine}_CoilId", "VARCHAR(50)", "PLC->API; Tablet->MES"),
        ("{Machine}_DowntimeReason", "VARCHAR(100)", "Tablet->MES"),
        ("{Machine}_DowntimeEvent", "VARCHAR(50)", "Tablet->MES; PLC->API"),
        ("Downtime_Event", "VARCHAR(50)", "Tablet->MES; PLC->API"),
        ("Event_Code", "VARCHAR(50)", "Tablet->MES; PLC->API"),
        ("{Machine}_OFF", "VARCHAR(50)", "PLC->API"),
        ("{Machine}_Performance", "DECIMAL(5,2)", "MES"),
        ("{Machine}_Quality", "DECIMAL(5,2)", "MES"),
        ("{Machine}_ForcedCode", "VARCHAR(50)", "Tablet->MES; PLC->API"),
        ("Q.Dw.Status", "VARCHAR(10)", "Tablet->MES"),
        ("Q_Drw_Status", "VARCHAR(10)", "Tablet->MES"),
        ("Q_Drw_WDia", "DECIMAL(12,4)", "Tablet->MES; PLC->API"),
        ("Q.Dw.WDim", "DECIMAL(12,4)", "Tablet->MES"),
        ("Q_Drw_WDim", "DECIMAL(12,4)", "Tablet->MES"),
        ("Q.Dw.DDir", "DECIMAL(12,4)", "Tablet->MES"),
        ("Q_Drw_DDir", "DECIMAL(12,4)", "Tablet->MES"),
        ("Q.Dw.WBreak", "INTEGER", "PLC->API; Tablet->MES"),
        ("Q_Drw_WBreak", "INTEGER", "PLC->API; Tablet->MES"),
        ("Q.Dw.Rds20", "DECIMAL(12,4)", "Tablet->MES"),
        ("Q_Drw_Rds20", "DECIMAL(12,4)", "Tablet->MES"),
        ("Q.Dw.Tens", "DECIMAL(12,4)", "Tablet->MES"),
        ("Q_Drw_Tens", "DECIMAL(12,4)", "Tablet->MES"),
        ("Q.Dw.Elong", "DECIMAL(12,4)", "Tablet->MES"),
        ("Q_Drw_Elong", "DECIMAL(12,4)", "Tablet->MES"),
        ("T.Dw.AreaRed", "DECIMAL(12,4)", "Tablet->MES; PLC->API"),
        ("T.Dw.OilID", "VARCHAR(50)", "Tablet->MES; PLC->API"),
        ("{Machine}_ID", "VARCHAR(50)", "MES"),
        ("{Machine}_Type", "VARCHAR(50)", "MES"),
        ("{Machine}_ProductionDate", "DATE", "MES; SAP->MES"),
        ("{Machine}_Workcenter", "VARCHAR(50)", "SAP->MES"),
        ("{Machine}_MaterialNumber", "VARCHAR(50)", "SAP->MES"),
        ("{Machine}_Unit", "VARCHAR(20)", "SAP->MES"),
        ("{Machine}_OEE", "DECIMAL(5,2)", "MES"),
        ("{Machine}_Availability", "DECIMAL(5,2)", "MES"),
        ("Order_SalesOrderId", "VARCHAR(100)", "SAP->MES"),
        ("Order_CustomerName", "VARCHAR(255)", "SAP->MES"),
        ("Order_LineItem", "VARCHAR(50)", "SAP->MES"),
        ("Order_ProductionOrderYear", "INTEGER", "SAP->MES"),
        ("Order_ProductionProposalId", "VARCHAR(100)", "SAP->MES"),
        ("Order_MaterialConsumed", "DECIMAL(12,3)", "MES->SAP"),
        # Extra from Excel / doc
        ("{Machine}_Speed.Optimized", "DECIMAL(10,2)", "PLC->API; Tablet->MES"),
        ("{Machine}_Speed_Optimized", "DECIMAL(10,2)", "PLC->API; Tablet->MES"),
        ("Q_Drw_DiaDev", "DECIMAL(12,4)", "Tablet->MES; MES"),
        ("Q_Drw_Rdc20", "DECIMAL(12,4)", "Tablet->MES"),
        ("Q.Dw.Rdc20", "DECIMAL(12,4)", "Tablet->MES"),
        ("T.Dw.OilLvl", "VARCHAR(50)", "Tablet->MES; PLC->API"),
        ("T.Dw.OilID", "VARCHAR(50)", "Tablet->MES; PLC->API"),
        # Section 10 - De xuat bo sung (Tablet / MES / SAP)
        ("Order_SapOrderId", "VARCHAR(100)", "SAP->MES"),
        ("Order_ConfirmationStatus", "VARCHAR(50)", "MES->SAP; SAP->MES"),
        ("Order_SyncToSapAt", "TIMESTAMPTZ", "MES"),
        ("Order_SyncToSapStatus", "VARCHAR(50)", "MES"),
        ("Order_ScrapToSap", "BOOLEAN/TEXT", "MES->SAP"),
        ("Order_MaterialBom", "TEXT/JSON", "SAP->MES"),
        ("{Order}_MaterialBom", "TEXT/JSON", "SAP->MES"),
        ("{Machine}_CurrentShiftStart", "TIMESTAMPTZ", "MES"),
        ("{Machine}_CurrentShiftEnd", "TIMESTAMPTZ", "MES"),
        ("{Machine}_CoilLength", "DECIMAL(12,2)", "PLC->API; MES"),
        ("{Machine}_CoilWeight", "DECIMAL(12,3)", "PLC->API; Tablet->MES"),
        ("Downtime_Reason", "VARCHAR(100)", "Tablet->MES"),
        ("event_time", "TIMESTAMPTZ", "Tablet->MES; PLC->API"),
        ("entered_by", "VARCHAR(255)", "Tablet->MES"),
    ]
    for tag, typ, src in pairs:
        d[norm(tag)] = (typ, src)
    return d

def main():
    if not EXCEL_PATH.exists():
        print("File not found:", EXCEL_PATH)
        return
    wb_src = openpyxl.load_workbook(EXCEL_PATH)
    ws_src = wb_src.active
    lookup = build_lookup()

    # Find header row (row where col B has "Name" or col C has "Struct")
    header_row = 1
    for r in range(1, min(8, ws_src.max_row + 1)):
        cell_b = ws_src.cell(row=r, column=2).value
        cell_c = ws_src.cell(row=r, column=3).value
        if cell_b and ("Name" in str(cell_b) or (cell_c and "Struct" in str(cell_c))):
            header_row = r
            break

    # Build output with exactly 6 columns: A-D from source, E=Kiểu, F=Nguồn
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    if ws_src.title:
        ws.title = ws_src.title[:31]

    for row in range(1, ws_src.max_row + 1):
        for col in range(1, 5):
            ws.cell(row=row, column=col, value=ws_src.cell(row=row, column=col).value)
        if row == header_row:
            ws.cell(row=row, column=5, value="Ki\u1ec3u d\u1eef li\u1ec7u")
            ws.cell(row=row, column=6, value="Ngu\u1ed3n d\u1eef li\u1ec7u")
        else:
            tag_cell = ws_src.cell(row=row, column=2).value
            typ, src = "", ""
            if tag_cell:
                raw = str(tag_cell).strip()
                if not raw.startswith("//") and raw and raw != "Name":
                    key = norm(raw)
                    if key:
                        typ, src = lookup.get(key, ("", ""))
                        if not typ and not src:
                            for k, v in lookup.items():
                                if k in key or key in k:
                                    typ, src = v
                                    break
            ws.cell(row=row, column=5, value=typ)
            ws.cell(row=row, column=6, value=src)

    # Append dòng đề xuất bổ sung (Tablet/MES/SAP) - tô màu vàng
    next_row = ws_src.max_row + 1
    ws.cell(row=next_row, column=2, value="// \u0110\u1ec1 xu\u1ea5t b\u1ed5 sung (Tablet / MES / SAP)")
    ws.cell(row=next_row, column=3, value="Struct")
    ws.cell(row=next_row, column=4, value="H\u1ea1ng m\u1ee5c \u0111\u1ec1 xu\u1ea5t th\u00eam")
    for c in range(1, 7):
        ws.cell(row=next_row, column=c).fill = YELLOW_FILL
    next_row += 1
    for tagname_b, mo_ta_d, typ, src in NEW_SUGGESTED_ROWS:
        ws.cell(row=next_row, column=2, value=tagname_b)
        ws.cell(row=next_row, column=4, value=mo_ta_d)
        ws.cell(row=next_row, column=5, value=typ)
        ws.cell(row=next_row, column=6, value=src)
        for c in range(1, 7):
            ws.cell(row=next_row, column=c).fill = YELLOW_FILL
        next_row += 1

    wb_out.save(OUTPUT_PATH)
    wb_src.close()
    print("Done. New file (6 columns):", str(OUTPUT_PATH))

if __name__ == "__main__":
    main()
